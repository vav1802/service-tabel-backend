from fastapi import FastAPI
from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl

app = FastAPI()

@app.get("/")
def root():
    return {"status": "Server is running"}

@app.post("/create-excel")
async def create_excel(payload: dict):
    data = payload.get("text", "")

    # ====== ТУТ ПОКА ПРОСТО ТЕСТ ======
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Полученные данные:"
    ws["A2"] = data

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=result.xlsx"}
    )
